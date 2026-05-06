import io
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from auth import get_current_user
from database import get_session
from models import Criteria, Project, User, Vote


router = APIRouter(prefix="/api/pipeline", tags=["pipeline"])


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _norm_title(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).lower()


def _cell(row: tuple[Any, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in {"1", "true", "yes", "y", "да", "истина"}:
        return True
    if s in {"0", "false", "no", "n", "нет", "ложь"}:
        return False
    return None


@router.post("/projects/import")
async def import_projects_xlsx(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .xlsx file")

    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid xlsx file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="XLSX is empty")

    header = [_norm_header(v) for v in rows[0]]
    if not any(header):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Header row is empty")

    # Fix the structure for initiatives_bank.xlsx.
    required_headers = {
        "название инициативы",
        "описание",
        "автор",
        "категория",
        "ожидаемый эффект",
        "срок реализации",
        "статус",
    }
    present_headers = {h for h in header if h}
    missing = sorted(required_headers - present_headers)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"XLSX has invalid structure. Missing columns: {', '.join(missing)}",
        )
    header_index = {h: i for i, h in enumerate(header)}
    title_idx = header_index.get("название инициативы")
    desc_idx = header_index.get("описание")
    author_idx = header_index.get("автор")
    category_idx = header_index.get("категория")
    effect_idx = header_index.get("ожидаемый эффект")
    timeline_idx = header_index.get("срок реализации")
    status_idx = header_index.get("статус")
    active_idx = header_index.get("активный")

    inserted = 0
    skipped = 0
    last_active_project: Project | None = None

    for r in rows[1:]:
        title = _cell(r, title_idx)
        if not _norm_title(title):
            skipped += 1
            continue

        desc = _cell(r, desc_idx)
        author = _cell(r, author_idx)
        category = _cell(r, category_idx)
        effect = _cell(r, effect_idx)
        timeline = _cell(r, timeline_idx)
        status_text = _cell(r, status_idx)
        full_description = "\n".join(
            [
                part
                for part in [
                    desc,
                    f"Автор: {author}" if author else "",
                    f"Категория: {category}" if category else "",
                    f"Ожидаемый эффект: {effect}" if effect else "",
                    f"Срок реализации: {timeline}" if timeline else "",
                    f"Статус: {status_text}" if status_text else "",
                ]
                if part
            ]
        )

        is_active = None
        if active_idx is not None and active_idx < len(r):
            is_active = _parse_bool(r[active_idx])

        project = Project(
            title=title,
            description=full_description or desc or "",
            is_active=bool(is_active) if is_active is not None else False,
        )
        session.add(project)
        inserted += 1
        if is_active is True:
            last_active_project = project

    await session.flush()

    # Ensure only one project is active (last "active" in the file wins).
    if last_active_project is not None:
        await session.execute(update(Project).values(is_active=False))
        last_active_project.is_active = True

    await session.commit()

    return {
        "inserted": inserted,
        "skipped": skipped,
        "active_set": bool(last_active_project is not None),
        "by_user": current_user.username,
    }


@router.get("/votes/export.xlsx")
async def export_votes_xlsx(
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    votes_res = await session.execute(
        select(Vote, User, Project, Criteria)
        .join(User, Vote.user_id == User.id)
        .join(Project, Vote.project_id == Project.id)
        .join(Criteria, Vote.criteria_id == Criteria.id)
        .order_by(Vote.created_at.asc(), Vote.id.asc())
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "votes"
    ws.append(
        [
            "vote_id",
            "created_at",
            "username",
            "display_name",
            "project_id",
            "project_title",
            "criteria_id",
            "criteria_name",
            "score",
        ]
    )

    for vote, user, project, criteria in votes_res.all():
        ws.append(
            [
                vote.id,
                vote.created_at.isoformat(sep=" ") if isinstance(vote.created_at, datetime) else str(vote.created_at),
                user.username,
                user.display_name,
                project.id,
                project.title,
                criteria.id,
                criteria.name,
                vote.score,
            ]
        )

    # Also export projects snapshot.
    ws_projects = wb.create_sheet("projects")
    ws_projects.append(["project_id", "title", "description", "is_active"])
    projects_res = await session.execute(select(Project).order_by(Project.id))
    for p in projects_res.scalars().all():
        ws_projects.append([p.id, p.title, p.description, bool(p.is_active)])

    # And criteria snapshot.
    ws_criteria = wb.create_sheet("criteria")
    ws_criteria.append(["criteria_id", "name", "max_score"])
    criteria_res = await session.execute(select(Criteria).order_by(Criteria.id))
    for c in criteria_res.scalars().all():
        ws_criteria.append([c.id, c.name, c.max_score])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now = datetime.utcnow().strftime("%Y-%m-%d_%H-%M")
    filename = f"votes_export_{now}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

